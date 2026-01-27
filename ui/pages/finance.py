"""
Finance Page - Finans Yönetimi
Gelir/gider takibi, raporlama, grafik analizi
"""

import flet as ft
from datetime import datetime, timedelta
from database.db_manager import DatabaseManager
from database.models import Transaction
from utils.logger import app_logger


class FinancePage:
    def __init__(self, page: ft.Page, db: DatabaseManager):
        self.page = page
        self.db = db
        
        # Filtreler
        self.selected_period = "month"  # today, week, month, year
        self.selected_type = "all"  # all, income, expense
        
        # UI Components
        self.summary_cards = ft.Row(spacing=20, wrap=True)
        self.transactions_table = ft.DataTable(
            columns=[
                ft.DataColumn(ft.Text("TARİH", size=11, weight="bold")),
                ft.DataColumn(ft.Text("TÜR", size=11, weight="bold")),
                ft.DataColumn(ft.Text("KATEGORİ", size=11, weight="bold")),
                ft.DataColumn(ft.Text("AÇIKLAMA", size=11, weight="bold")),
                ft.DataColumn(ft.Text("TUTAR", size=11, weight="bold"), numeric=True),
                ft.DataColumn(ft.Text("İŞLEM", size=11, weight="bold")),
            ],
            heading_row_color="#f8f9fa",
            width=float("inf")
        )
        
        self.chart_container = ft.Container()
        
        # Yeni işlem form alanları
        self.dd_type = ft.Dropdown(
            label="İşlem Türü *",
            options=[
                ft.dropdown.Option("Gelir"),
                ft.dropdown.Option("Gider")
            ],
            value="Gelir",
            width=200
        )
        
        self.dd_category = ft.Dropdown(
            label="Kategori *",
            width=200,
            on_change=self.on_type_changed
        )
        
        self.txt_amount = ft.TextField(
            label="Tutar *",
            suffix_text="₺",
            keyboard_type=ft.KeyboardType.NUMBER,
            width=200
        )
        
        self.txt_description = ft.TextField(
            label="Açıklama",
            width=400
        )
        
        self.date_picker = ft.DatePicker(
            on_change=self.on_date_picked
        )
        self.page.overlay.append(self.date_picker)
        
        self.selected_transaction_date = datetime.now()
        self.txt_date_display = ft.Text(
            self.selected_transaction_date.strftime("%d.%m.%Y")
        )
        
        self.load_categories()
        
    def view(self):
        """Ana görünüm"""
        self.load_data()
        
        # Header
        header = ft.Container(
            content=ft.Row([
                ft.Icon(ft.Icons.ACCOUNT_BALANCE_WALLET, color="teal", size=30),
                ft.Column([
                    ft.Text("Finans Yönetimi", size=24, weight="bold"),
                    ft.Text("Gelir ve gider takibi", size=12, color="grey")
                ], spacing=0),
                ft.Container(expand=True),
                ft.Row([
                    ft.ElevatedButton(
                        "Yeni İşlem",
                        icon=ft.Icons.ADD,
                        bgcolor="teal",
                        color="white",
                        on_click=self.open_new_transaction_dialog
                    )
                ])
            ]),
            padding=20,
            bgcolor="white",
            border_radius=15
        )
        
        # Summary Cards
        summary_section = ft.Container(
            content=ft.Column([
                ft.Row([
                    ft.Text("Finansal Özet", weight="bold"),
                    ft.Container(expand=True),
                    ft.SegmentedButton(
                        selected={"month"},
                        allow_empty_selection=False,
                        allow_multiple_selection=False,
                        segments=[
                            ft.Segment(
                                value="today",
                                label=ft.Text("Bugün"),
                                icon=ft.Icon(ft.Icons.TODAY)
                            ),
                            ft.Segment(
                                value="week",
                                label=ft.Text("Hafta"),
                                icon=ft.Icon(ft.Icons.DATE_RANGE)
                            ),
                            ft.Segment(
                                value="month",
                                label=ft.Text("Ay"),
                                icon=ft.Icon(ft.Icons.CALENDAR_MONTH)
                            ),
                            ft.Segment(
                                value="year",
                                label=ft.Text("Yıl"),
                                icon=ft.Icon(ft.Icons.CALENDAR_TODAY)
                            )
                        ],
                        on_change=self.on_period_changed
                    )
                ]),
                ft.Divider(height=10, color="transparent"),
                self.summary_cards
            ], spacing=10),
            padding=20,
            bgcolor="white",
            border_radius=15
        )
        
        # Chart
        chart_section = ft.Container(
            content=ft.Column([
                ft.Text("Gelir/Gider Trendi", weight="bold"),
                self.chart_container
            ]),
            padding=20,
            bgcolor="white",
            border_radius=15
        )
        
        # Transactions Table
        table_section = ft.Container(
            content=ft.Column([
                ft.Row([
                    ft.Text("İşlem Geçmişi", size=16, weight="bold"),
                    ft.Container(expand=True),
                    ft.Dropdown(
                        label="Filtre",
                        options=[
                            ft.dropdown.Option("all", "Tümü"),
                            ft.dropdown.Option("income", "Gelir"),
                            ft.dropdown.Option("expense", "Gider")
                        ],
                        value="all",
                        width=150,
                        on_change=self.on_type_filter_changed
                    ),
                    ft.IconButton(
                        ft.Icons.DOWNLOAD,
                        tooltip="Excel'e Aktar",
                        on_click=self.export_to_excel
                    )
                ]),
                ft.Divider(),
                ft.Container(
                    content=ft.Column([self.transactions_table], scroll=ft.ScrollMode.AUTO),
                    height=400
                )
            ]),
            padding=20,
            bgcolor="white",
            border_radius=15,
            expand=True
        )
        
        return ft.View(
            "/finance",
            controls=[
                ft.Container(
                    content=ft.Column([
                        header,
                        summary_section,
                        chart_section,
                        table_section
                    ], spacing=15, expand=True),
                    padding=30,
                    bgcolor="#f8f9fa",
                    expand=True
                )
            ],
            padding=0
        )
    
    def load_data(self):
        """Tüm verileri yükle"""
        try:
            self.load_summary()
            self.load_chart()
            self.load_transactions()
            
        except Exception as e:
            app_logger.error(f"Finance data loading error: {e}")
    
    def load_summary(self):
        """Özet kartları yükle"""
        try:
            # Tarih aralığı belirle
            end_date = datetime.now()
            
            if self.selected_period == "today":
                start_date = end_date.replace(hour=0, minute=0, second=0)
            elif self.selected_period == "week":
                start_date = end_date - timedelta(days=7)
            elif self.selected_period == "month":
                start_date = end_date.replace(day=1)
            else:  # year
                start_date = end_date.replace(month=1, day=1)
            
            # Verileri çek
            total_income = self.db.get_total_income(start_date, end_date)
            total_expense = self.db.get_total_expense(start_date, end_date)
            net_balance = total_income - total_expense
            
            # Önceki dönemle karşılaştırma
            period_diff = end_date - start_date
            prev_start = start_date - period_diff
            prev_end = start_date
            
            prev_income = self.db.get_total_income(prev_start, prev_end)
            income_change = ((total_income - prev_income) / prev_income * 100) if prev_income > 0 else 0
            
            self.summary_cards.controls = [
                self._summary_card(
                    "Toplam Gelir",
                    f"₺{total_income:,.2f}",
                    f"%{income_change:+.1f} önceki döneme göre",
                    ft.Icons.ARROW_UPWARD,
                    "green"
                ),
                self._summary_card(
                    "Toplam Gider",
                    f"₺{total_expense:,.2f}",
                    f"{self.db.get_transaction_count(start_date, end_date, 'Gider')} işlem",
                    ft.Icons.ARROW_DOWNWARD,
                    "red"
                ),
                self._summary_card(
                    "Net Durum",
                    f"₺{net_balance:,.2f}",
                    "Gelir - Gider",
                    ft.Icons.ACCOUNT_BALANCE,
                    "blue" if net_balance >= 0 else "orange"
                )
            ]
            
        except Exception as e:
            app_logger.error(f"Load summary error: {e}")
    
    def _summary_card(self, title, value, subtitle, icon, color):
        """Özet kartı"""
        return ft.Container(
            content=ft.Column([
                ft.Row([
                    ft.Container(
                        content=ft.Icon(icon, color=color, size=24),
                        padding=12,
                        bgcolor=ft.Colors.with_opacity(0.1, color),
                        border_radius=12
                    ),
                    ft.Container(expand=True)
                ]),
                ft.Text(value, size=28, weight="bold", color=color),
                ft.Text(title, size=14, color="grey"),
                ft.Text(subtitle, size=12, color="grey")
            ], spacing=8),
            padding=20,
            bgcolor="white",
            border_radius=15,
            border=ft.border.all(1, "#f0f0f0"),
            shadow=ft.BoxShadow(
                blur_radius=5,
                color=ft.Colors.with_opacity(0.05, "black")
            ),
            width=280
        )
    
    def load_chart(self):
        """Grafik yükle"""
        try:
            # Son 30 günün verileri
            end_date = datetime.now()
            start_date = end_date - timedelta(days=30)
            
            daily_data = self.db.get_daily_financial_summary(start_date, end_date)
            
            if not daily_data:
                self.chart_container.content = ft.Text("Veri yok", color="grey")
                return
            
            # İki seri: Gelir ve Gider
            income_points = []
            expense_points = []
            
            for i, (date, income, expense) in enumerate(daily_data):
                income_points.append(ft.LineChartDataPoint(i, income))
                expense_points.append(ft.LineChartDataPoint(i, expense))
            
            chart = ft.LineChart(
                data_series=[
                    ft.LineChartData(
                        data_points=income_points,
                        stroke_width=3,
                        color="green",
                        curved=True,
                        stroke_cap_round=True,
                    ),
                    ft.LineChartData(
                        data_points=expense_points,
                        stroke_width=3,
                        color="red",
                        curved=True,
                        stroke_cap_round=True,
                    )
                ],
                border=ft.border.all(1, ft.Colors.GREY_200),
                left_axis=ft.ChartAxis(
                    labels_size=40,
                    title=ft.Text("Tutar (₺)")
                ),
                bottom_axis=ft.ChartAxis(
                    labels=[
                        ft.ChartAxisLabel(
                            value=i,
                            label=ft.Text(
                                daily_data[i][0].strftime("%d"),
                                size=10
                            )
                        ) for i in range(0, len(daily_data), 5)
                    ]
                ),
                horizontal_grid_lines=ft.ChartGridLines(
                    color=ft.Colors.GREY_100,
                    width=1
                ),
                expand=True
            )
            
            self.chart_container.content = ft.Container(
                content=chart,
                height=300
            )
            
        except Exception as e:
            app_logger.error(f"Load chart error: {e}")
    
    def load_transactions(self):
        """İşlemleri yükle"""
        try:
            self.transactions_table.rows.clear()
            
            # Filtreye göre işlemleri çek
            if self.selected_type == "all":
                transactions = self.db.get_all_transactions(limit=100)
            elif self.selected_type == "income":
                transactions = self.db.get_transactions_by_type("Gelir", limit=100)
            else:
                transactions = self.db.get_transactions_by_type("Gider", limit=100)
            
            for trans in transactions:
                # Renk kodlama
                is_income = (trans.type == "Gelir")
                amount_color = "green" if is_income else "red"
                icon = ft.Icons.ARROW_UPWARD if is_income else ft.Icons.ARROW_DOWNWARD
                
                self.transactions_table.rows.append(
                    ft.DataRow(cells=[
                        ft.DataCell(
                            ft.Text(
                                trans.date.strftime("%d.%m.%Y") if isinstance(trans.date, datetime) else trans.date,
                                size=12
                            )
                        ),
                        ft.DataCell(
                            ft.Container(
                                content=ft.Row([
                                    ft.Icon(icon, size=12, color=amount_color),
                                    ft.Text(trans.type, color=amount_color, size=12)
                                ], spacing=5),
                                bgcolor=ft.Colors.with_opacity(0.1, amount_color),
                                padding=5,
                                border_radius=5
                            )
                        ),
                        ft.DataCell(ft.Text(trans.category or "-", size=12)),
                        ft.DataCell(ft.Text(trans.description or "-", size=12)),
                        ft.DataCell(
                            ft.Text(
                                f"₺{trans.amount:,.2f}",
                                weight="bold",
                                color=amount_color,
                                size=13
                            )
                        ),
                        ft.DataCell(
                            ft.IconButton(
                                ft.Icons.DELETE_OUTLINE,
                                icon_color="red",
                                icon_size=18,
                                tooltip="Sil",
                                on_click=lambda _, tid=trans.id: self.delete_transaction(tid)
                            )
                        )
                    ])
                )
            
            if not transactions:
                self.transactions_table.rows.append(
                    ft.DataRow(cells=[
                        ft.DataCell(
                            ft.Text("İşlem bulunamadı", italic=True, color="grey")
                        ),
                        ft.DataCell(ft.Text("")),  # Empty cell
                        ft.DataCell(ft.Text("")),  # Empty cell
                        ft.DataCell(ft.Text("")),  # Empty cell
                        ft.DataCell(ft.Text("")),  # Empty cell
                        ft.DataCell(ft.Text("")),  # Empty cell
                    ])
                )
            
            self.transactions_table.update()
            
        except Exception as e:
            app_logger.error(f"Load transactions error: {e}")
    
    def load_categories(self):
        """Kategorileri yükle"""
        income_categories = [
            "Muayene",
            "Tedavi",
            "Konsültasyon",
            "Rapor",
            "Diğer"
        ]
        
        expense_categories = [
            "Kira",
            "Maaş",
            "Malzeme",
            "Elektrik/Su",
            "Bakım/Onarım",
            "Vergi",
            "Diğer"
        ]
        
        # İlk yüklemede gelir kategorileri
        self.dd_category.options = [
            ft.dropdown.Option(cat) for cat in income_categories
        ]
    
    def on_type_changed(self, e):
        """İşlem türü değiştiğinde kategorileri güncelle"""
        if self.dd_type.value == "Gelir":
            categories = [
                "Muayene",
                "Tedavi",
                "Konsültasyon",
                "Rapor",
                "Diğer"
            ]
        else:
            categories = [
                "Kira",
                "Maaş",
                "Malzeme",
                "Elektrik/Su",
                "Bakım/Onarım",
                "Vergi",
                "Diğer"
            ]
        
        self.dd_category.options = [
            ft.dropdown.Option(cat) for cat in categories
        ]
        self.dd_category.value = None
        self.dd_category.update()
    
    def on_period_changed(self, e):
        """Dönem filtresi değiştiğinde"""
        self.selected_period = list(e.control.selected)[0]
        self.load_summary()
        self.summary_cards.update()
    
    def on_type_filter_changed(self, e):
        """Tür filtresi değiştiğinde"""
        self.selected_type = e.control.value
        self.load_transactions()
    
    def open_new_transaction_dialog(self, e):
        """Yeni işlem dialogu"""
        # Form sıfırla
        self.dd_type.value = "Gelir"
        self.dd_category.value = None
        self.txt_amount.value = ""
        self.txt_description.value = ""
        self.selected_transaction_date = datetime.now()
        self.txt_date_display.value = self.selected_transaction_date.strftime("%d.%m.%Y")
        self.load_categories()
        
        dialog = ft.AlertDialog(
            title=ft.Text("Yeni İşlem Ekle"),
            content=ft.Container(
                content=ft.Column([
                    self.dd_type,
                    self.dd_category,
                    self.txt_amount,
                    self.txt_description,
                    ft.Row([
                        ft.Text("Tarih:", weight="bold"),
                        self.txt_date_display,
                        ft.IconButton(
                            ft.Icons.CALENDAR_MONTH,
                            on_click=lambda _: self.date_picker.pick_date()
                        )
                    ])
                ], tight=True, scroll=ft.ScrollMode.AUTO),
                width=400,
                height=350
            ),
            actions=[
                ft.TextButton("İptal", on_click=lambda _: self.page.close(dialog)),
                ft.ElevatedButton(
                    "Kaydet",
                    icon=ft.Icons.SAVE,
                    bgcolor="teal",
                    color="white",
                    on_click=lambda _: self.save_transaction(dialog)
                )
            ]
        )
        
        self.page.open(dialog)
    
    def on_date_picked(self, e):
        """Tarih seçildiğinde"""
        if self.date_picker.value:
            self.selected_transaction_date = self.date_picker.value
            self.txt_date_display.value = self.selected_transaction_date.strftime("%d.%m.%Y")
            self.txt_date_display.update()
    
    def save_transaction(self, dialog):
        """İşlemi kaydet"""
        try:
            # Validasyon
            if not self.dd_type.value:
                self.page.open(ft.SnackBar(
                    ft.Text("Lütfen işlem türü seçin"),
                    bgcolor="red"
                ))
                return
            
            if not self.dd_category.value:
                self.page.open(ft.SnackBar(
                    ft.Text("Lütfen kategori seçin"),
                    bgcolor="red"
                ))
                return
            
            if not self.txt_amount.value:
                self.page.open(ft.SnackBar(
                    ft.Text("Lütfen tutar girin"),
                    bgcolor="red"
                ))
                return
            
            try:
                amount = float(self.txt_amount.value.replace(",", "."))
            except ValueError:
                self.page.open(ft.SnackBar(
                    ft.Text("Geçersiz tutar formatı"),
                    bgcolor="red"
                ))
                return
            
            # Transaction objesi oluştur
            transaction = Transaction(
                id=None,
                type=self.dd_type.value,
                category=self.dd_category.value,
                amount=amount,
                description=self.txt_description.value,
                date=self.selected_transaction_date
            )
            
            # Veritabanına kaydet
            trans_id = self.db.add_transaction(transaction)
            
            # Audit log
            self.db.add_audit_log(
                user_id=self.page.session.get("user_id"),
                action_type="financial",
                description=f"{transaction.type} işlemi eklendi: ₺{amount:,.2f}",
                ip_address=self.page.session.get("ip_address")
            )
            
            # Dialog kapat ve verileri yenile
            self.page.close(dialog)
            self.load_data()
            
            self.page.open(ft.SnackBar(
                ft.Text("✅ İşlem başarıyla kaydedildi"),
                bgcolor="green"
            ))
            
        except Exception as e:
            app_logger.error(f"Save transaction error: {e}")
            self.page.open(ft.SnackBar(
                ft.Text(f"❌ Kayıt hatası: {e}"),
                bgcolor="red"
            ))
    
    def delete_transaction(self, transaction_id):
        """İşlemi sil"""
        def confirm_delete(e):
            try:
                self.db.delete_transaction(transaction_id)
                
                self.page.close(dialog)
                self.load_data()
                
                self.page.open(ft.SnackBar(
                    ft.Text("İşlem silindi"),
                    bgcolor="green"
                ))
                
            except Exception as ex:
                app_logger.error(f"Delete transaction error: {ex}")
                self.page.open(ft.SnackBar(
                    ft.Text(f"Silme hatası: {ex}"),
                    bgcolor="red"
                ))
        
        dialog = ft.AlertDialog(
            title=ft.Text("İşlemi Sil"),
            content=ft.Text("Bu işlemi silmek istediğinizden emin misiniz?"),
            actions=[
                ft.TextButton("İptal", on_click=lambda _: self.page.close(dialog)),
                ft.ElevatedButton(
                    "Sil",
                    bgcolor="red",
                    color="white",
                    on_click=confirm_delete
                )
            ]
        )
        
        self.page.open(dialog)
    
    def export_to_excel(self, e):
        """Excel'e aktar"""
        try:
            import openpyxl
            from datetime import datetime
            
            # Workbook oluştur
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Finansal Rapor"
            
            # Header
            headers = ["Tarih", "Tür", "Kategori", "Açıklama", "Tutar (₺)"]
            ws.append(headers)
            
            # Stil
            for cell in ws[1]:
                cell.font = openpyxl.styles.Font(bold=True)
                cell.fill = openpyxl.styles.PatternFill(
                    start_color="009688",
                    end_color="009688",
                    fill_type="solid"
                )
                cell.font = openpyxl.styles.Font(color="FFFFFF", bold=True)
            
            # Veri
            transactions = self.db.get_all_transactions(limit=1000)
            for trans in transactions:
                date_str = trans.date.strftime("%d.%m.%Y") if isinstance(trans.date, datetime) else trans.date
                ws.append([
                    date_str,
                    trans.type,
                    trans.category,
                    trans.description,
                    trans.amount
                ])
            
            # Kaydet
            filename = f"finans_raporu_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            wb.save(filename)
            
            self.page.open(ft.SnackBar(
                ft.Text(f"Rapor oluşturuldu: {filename}"),
                bgcolor="green"
            ))
            
            # Dosyayı aç
            import os
            import subprocess
            if os.name == 'nt':  # Windows
                os.startfile(filename)
            else:  # Mac/Linux
                subprocess.call(('open', filename))
            
        except Exception as ex:
            app_logger.error(f"Export error: {ex}")
            self.page.open(ft.SnackBar(
                ft.Text(f"Dışa aktarma hatası: {ex}"),
                bgcolor="red"
            ))